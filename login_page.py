
from tkinter import *
from tkinter import ttk
from tkcalendar import DateEntry
from tkinter import messagebox
import sqlite3  as db
import docx
import openpyxl
#import win32api

def establishconnect():
    connectobject = db.connect("shopManagement.db")
    c = connectobject.cursor()
    sql = '''
    create table if not exists sell (
        customerName string,
        date string,
        product string,
        price number,
        quantity number,
        total number
        )
    '''
    c.execute(sql)
    connectobject.commit()   

establishconnect()    
mainwindow=Tk()
mainwindow.title("store management project")
tab = ttk.Notebook(mainwindow) 
window1= ttk.Frame(tab)
window2=ttk.Frame(tab)

tab.add(window1, text ='stock') 
tab.add(window2, text ='sell') 
tab.pack(expand = 1, fill ="both") 

def connection2():
    connectobject2 = db.connect("shopManagement.db")
    c = connectobject2.cursor()
    sql = '''
    create table if not exists stock (
        date string,
        product string,
        price number,
        quantity number, 
        customerName string
        )
    '''
    c.execute(sql)
    connectobject2.commit()   

connection2() 

def Stock():
    global dateE2,quantity, name, price, customerName
    connectobject = db.connect("shopManagement.db")
    c = connectobject.cursor()  
    sql = '''
            INSERT INTO stock VALUES 
            (?, ?, ?, ?, ?)
            '''
    c.execute(sql, (dateE2.get(),name.get(),price.get(),quantity.get(), customerName.get()))
    connectobject.commit()  
    connectobject.close()

def viewingStock():
    connectobject = db.connect("shopManagement.db")
    c = connectobject.cursor()  

    sql = 'Select * from stock'
    c.execute(sql)

    rows=c.fetchall()
    viewingarea2.config(state='normal')
    viewingarea2.insert(END, f"   Date \t     Product\t      Price\t     \t qantity\t  customerName\t  \n")
    
    for i in rows:
        allrows=" "
        for j in i:
            allrows+=str(j)+'\t \t'
        allrows+='\n'
        viewingarea2.insert(END, allrows)
    viewingarea2.config(state='disabled')
    
def clear_text_area():
    viewingarea2.config(state='normal')
    viewingarea2.delete('1.0', END)
    viewingarea2.config(state='disabled')
    
clear = Button(window1, text="clear", command=clear_text_area,
font=('arial',15,'bold'), bg="blue",width=20)
clear.grid(row=7,column=0,padx=7,pady=7)

#add the print method here 

def printB():
    # Create a variable to store the text to be printed
    text_to_print = viewingarea2.get("1.0", "end-1c")

    # Create a new Word document
    doc = docx.Document()

    # Add the text to the document
    doc.add_paragraph(text_to_print)

    # Save the document as a .docx file
    doc.save("khemsafe.docx")

    # Print message to confirm that writing is complete
    print("Text has been printed to khemsafe.docx!")
    
    #working on this part
def updateeed():
    connectobject = db.connect("shopManagement.db")

# create a cursor object
    c = connectobject.cursor()  

# update a row in the table
    c.execute("UPDATE stock SET column1 = ?, column2 = ?, column3 = ?, column4 = ? WHERE id = ? ", ('new value 1', 'new value 2', 1))

# commit the changes
    connectobject.commit()

# close the cursor and connection
    c.close()

def export_to_excel():
    # create a new Workbook object
    wb = openpyxl.Workbook()

    # select the active worksheet
    ws = wb.active

    # get the contents of the text field
    text_to_print = viewingarea2.get("1.0", "end-1c")

    # split the text into rows based on newline characters
    rows = text_to_print.split("\n")

    # iterate over each row and write it to the worksheet
    for i, row in enumerate(rows):
        # split the row into columns based on tab characters
        cols = row.split("\t")

        # iterate over each column and write it to the cell
        for j, col in enumerate(cols):
            # convert the row and column indices to Excel-style cell references (e.g. A1)
            cell = openpyxl.utils.cell.get_column_letter(j+1) + str(i+1)

            # write the cell value to the worksheet
            ws[cell] = col

    # save the workbook to a file
    wb.save("khemsafe.xlsx")

    connectobject = db.connect("shopManagement.db")
    c = connectobject.cursor()  

    sql = 'Select * from stock'
    c.execute(sql)

    rows=c.fetchall()
    text_to_print += "  Date \t     Product\t      Price\t     \t qantity \t  customerName \t \n"
    
    for i in rows:
        allrows=" "
        for j in i:
            allrows+=str(j)+'\t \t'
        allrows+='\n'
        text_to_print += allrows

    # Print the text using the default system printer
    #win32api.ShellExecute(0, "print", text_to_print, None,  ".",  0)
printButton = Button(
    window1,
    text="Print",
    command=lambda: (printB(), export_to_excel()),
    font=('Arial', 15, 'bold'),
    bg="blue",
    width=20
)
printButton.grid(row=7, column=1, padx=7, pady=7)

dateL=Label(window1,text="Date",width=12,font=('arial',15,'bold'))
dateL.grid(row=0,column=0,padx=7,pady=7)

dateE2=DateEntry(window1,width=12,font=('arial',15,'bold'))
dateE2.grid(row=0,column=1,padx=7,pady=7)

l1=Label(window1, text="Product",font=('arial',15,'bold'),width=12)
l1.grid(row=1,column=0,padx=7,pady=7)

l1=Label(window1, text="Price",font=('arial',15,'bold'),width=12)
l1.grid(row=2,column=0,padx=7,pady=7)

l1=Label(window1, text="Quantity",font=('arial',15,'bold'),width=12)
l1.grid(row=3,column=0,padx=7,pady=7)

#work on the position on the screen
customerName =Label(window1,text=" Customer Name ",width=12,font=('arial',15,'bold'))
customerName.grid(row=4,column=0,padx=7,pady=7)

name=StringVar()
price=IntVar()
quantity=IntVar()
customerName = StringVar()

Name=Entry(window1,textvariable=name,font=('arial',15,'bold'),width=20)
Name.grid(row=1,column=1,padx=7,pady=7)

Price=Entry(window1,textvariable=price,font=('arial',15,'bold'),width=20)
Price.grid(row=2,column=1,padx=7,pady=7)
Price.delete(0, END)

Quantity=Entry(window1,textvariable=quantity,font=('arial',15,'bold'),width=20)
Quantity.grid(row=3,column=1,padx=7,pady=7)
Quantity.delete(0, END)

#work on the position
nameEntry=Entry(window1,width=20,textvariable = customerName,font=('arial',15,'bold'))
nameEntry.grid(row=4,column=1,padx=7,pady=7)
nameEntry.delete(0, END)

addbutton=Button(window1,command=Stock,text="Add",
font=('arial',15,'bold'),bg="blue",width=20)
addbutton.grid(row=5,column=1,padx=7,pady=7)

viewingarea2=Text(window1,width= 70 )
viewingarea2.grid(row=6,column=0,columnspan=2)

viewbutton2=Button(window1,command=viewingStock,text="View Stock",
font=('arial',15,'bold'),bg="blue",width=20 )

viewbutton2.grid(row=5,column=0,padx=7,pady=7)

def Bill():
    connectobject = db.connect("shopManagement.db")
    c = connectobject.cursor()  

    global areaforbill
    if p1quant.get()==0 and p2quant.get()==0 and p3quant.get()==0 and p4quant.get()==0:
        messagebox.showerror("Error","No product purchased")
    else:
        areaforbill.delete('1.0',END)
        areaforbill.insert(END,"\t|| store management project ||")
        areaforbill.insert(END,"\n_________________________________________\n")
        areaforbill.insert(END,"\nDate \t Price \tProducts \tQTY \t Total")
        areaforbill.insert(END,"\n===========================================")

        price= IntVar()
        price2=IntVar()
        price3=IntVar()
        price4=IntVar()

        print(datee.get())
        price=price2=price3=price4=0

        if p1quant.get()!=0:
            price=p1quant.get()*pricep1.get()
            print(price)
            areaforbill.insert(END,f"\n{datee.get()}\t Product-1 \t{pricep1.get()}\t {p1quant.get()}\t {price}")

            sql = '''
            INSERT INTO Sell VALUES 
            (?, ?, ?, ?,?)
            '''
            c.execute(sql,(datee.get(),'Product-1',pricep1.get(),p1quant.get(),price))
            connectobject.commit() 

        if p2quant.get()!=0:
            price2=(p2quant.get()*pricep2.get())
            print(price2)
            areaforbill.insert(END,f"\n{datee.get()}\t Product-2 \t{pricep2.get()}\t {p2quant.get()}\t {price2}")

            sql = '''
            INSERT INTO Sell VALUES 
            (?, ?, ?, ?,?)
            '''
            print(datee.get(),'Product-2',pricep2.get(),p2quant.get(),price2)
            c.execute(sql,(datee.get(),'Product-2',pricep2.get(),p2quant.get(),price2))
            connectobject.commit() 

        if p3quant.get()!=0:
            price3=p3quant.get()*pricep1.get()
            print(price3)
            areaforbill.insert(END,f"\n{datee.get()}\tProduct-3 \t{pricep3.get()}\t {p3quant.get()}\t {price3}")

            sql = '''
            INSERT INTO Sell VALUES 
            (?, ?, ?, ?,?)
            '''
            c.execute(sql,(datee.get(),'Product-3',pricep3.get(),p3quant.get(),price3))
            connectobject.commit() 

        if p4quant.get()!=0:
            price4=p4quant.get()*pricep1.get()
            areaforbill.insert(END,f"\n{datee.get()}\tProduct-4 \t{pricep4.get()}\t {p4quant.get()}\t {price4}")

            sql = '''
            INSERT INTO Sell VALUES 
            (?, ?, ?, ?,?)
            '''
            c.execute(sql,(datee.get(),'Product-4',pricep4.get(),p4quant.get(),price4))
            connectobject.commit() 

        Total=IntVar()
        Total=price+price2+price3+price4

        quantity=IntVar()
        quantity=p1quant.get()+p2quant.get()+p3quant.get()+p4quant.get()
        areaforbill.insert(END,f"\nTotal \t \t  \t{quantity}\t {Total}")

def view():
    connectobject = db.connect("shopManagement.db")
    c = connectobject.cursor()  

    sql = 'Select * from Sell'
    c.execute(sql)

    rows=c.fetchall()
    viewingarea.insert(END,f"Date\t Product\t  Price of 1\t  Quantity\t  Price\n")
    
    
    for i in rows:
        allrows=""
        for j in i:
            allrows+=str(j)+'\t'
        allrows+='\n'
        viewingarea.insert(END,allrows)

datel=Label(window1,text="Date",width=12,font=('arial',15,'bold'))
datel.grid(row=0,column=0,padx=7,pady=7)

datee=DateEntry(window1,width=12,font=('arial',15,'bold'))
datee.grid(row=0,column=1,padx=7,pady=7)

l1=Label(window1, text="Product Name",font=('arial', 15, 'bold'),width=12)
l1.grid(row=1,column=0,padx=7,pady=7)

namep1=StringVar()
namep1.set('Laptop')

pricep1=IntVar()
pricep1.set(80)

p1quant=IntVar()
p1quant.set(0)

l1=Label(window2, text=namep1.get(),font=('arial',15,'bold'),width=12)
l1.grid(row=2,column=0,padx=7,pady=7)

l1=Label(window2, text=pricep1.get(),font=('arial',15,'bold'),width=12)
l1.grid(row=2,column=1,padx=7,pady=7)

t1=Entry(window2,textvariable=p1quant,font=('arial',15,'bold'),width=12)
t1.grid(row=2,column=2,padx=7,pady=7)

namep2=StringVar()
namep2.set('Desktop')

pricep2=IntVar()
pricep2.set(20)

p2quant=IntVar()
p2quant.set(0)

l1=Label(window2, text=namep2.get(),font=('arial',15,'bold'),width=12)
l1.grid(row=3,column=0,padx=7,pady=7)

l1=Label(window2, text=pricep2.get(),font=('arial',15,'bold'),width=12)
l1.grid(row=3,column=1,padx=7,pady=7)

t1=Entry(window2,textvariable=p2quant,font=('arial',15,'bold'),width=12)
t1.grid(row=3,column=2,padx=7,pady=7)

namep3=StringVar()
namep3.set('Charger')

pricep3=IntVar()
pricep3.set(100)

p3quant=IntVar()
p3quant.set(0)

l1=Label(window2, text=namep3.get(),font=('arial',15,'bold'),width=12)
l1.grid(row=4,column=0,padx=7,pady=7)

l1=Label(window2, text=pricep3.get(),font=('arial',15,'bold'),width=12)
l1.grid(row=4,column=1,padx=7,pady=7)

t1=Entry(window2,textvariable=p3quant,font=('arial',15,'bold'),width=12)
t1.grid(row=4,column=2,padx=7,pady=7)

namep4=StringVar()
namep4.set('Laptop battery')

pricep4=IntVar()
pricep4.set(400)

p4quant=IntVar()
p4quant.set(0)

l1=Label(window2, text=namep4.get(),font=('arial',15,'bold'),width=12)
l1.grid(row=5,column=0,padx=7,pady=7)

l1=Label(window2, text=pricep4.get(),font=('arial',15,'bold'),width=12)
l1.grid(row=5,column=1,padx=7,pady=7)

t1=Entry(window2,textvariable=p4quant,font=('arial',15,'bold'),width=12)
t1.grid(row=5,column=2,padx=7,pady=7)


areaforbill=Text(window2)

submitbutton=Button(window2,command=Bill,text="Bill",
font=('arial',15,'bold'),bg="blue",width=20 )
submitbutton.grid(row=6,column=2,padx=7,pady=7)

viewbutton=Button(window2,command=view,text="View All Sellings",
font=('arial',  15,'bold'),bg="blue", width=20 )
viewbutton.grid(row=6,column=0,padx=7,pady=7)

areaforbill.grid(row=9,column=2)
viewingarea=Text(window2)
viewingarea.grid(row=9,column=0)

mainloop()